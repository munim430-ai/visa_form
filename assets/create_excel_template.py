"""
create_excel_template.py
------------------------
Generates assets/visa_template.xlsx — the input file operators fill
before uploading to the app.  Column order must match COLUMNS in
src/excel.js exactly.

Usage:
    pip install openpyxl
    python assets/create_excel_template.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = os.path.join(os.path.dirname(__file__), 'visa_template.xlsx')

# Column definitions — (excel_col, field_key, header_label, width, example_1, example_2)
# NOTE: stay_duration removed (hardcoded as "365 DAYS" in generate.js)
COLUMNS = [
    ('A',  'family_name',          'Family Name (English)',                  22, 'ALI',           'RAHMAN'),
    ('B',  'given_name',           'Given Name(s) (English)',                22, 'SHIAM',          'MOHAMMED'),
    ('C',  'korean_name',          '漢字姓名 (optional)',                     18, '',               ''),
    ('D',  'birth_year',           'Birth Year (YYYY)',                      14, '2003',           '1999'),
    ('E',  'birth_month',          'Birth Month (MM)',                       14, '01',             '07'),
    ('F',  'birth_day',            'Birth Day (DD)',                         12, '19',             '14'),
    ('G',  'gender',               'Gender (M/F)',                           10, 'M',              'F'),
    ('H',  'national_id',          'National ID No.',                        20, '2417484512',     '1234567890'),
    ('I',  'passport_number',      'Passport Number',                        18, 'A09596926',      'B12345678'),
    ('J',  'passport_issue_place', 'Passport Place of Issue',                22, 'DIP/DHAKA',      'DIP/DHAKA'),
    ('K',  'passport_issue_date',  'Passport Issue Date (YYYY-MM-DD)',       24, '2025-12-08',     '2020-03-15'),
    ('L',  'passport_expiry_date', 'Passport Expiry Date (YYYY-MM-DD)',      24, '2035-12-07',     '2030-03-14'),
    ('M',  'university_name',      'University / Institution in Korea',      34, 'MOKWON UNIVERSITY', 'HANYANG UNIVERSITY'),
    ('N',  'course_name',          'Course / Program Name',                  28, 'EAP PROGRAM',    'KOREAN LANGUAGE COURSE'),
    ('O',  'university_address',   'University Address',                     40, '대전광역시 서구 도안북로 88', '서울특별시 성동구 왕십리로 222'),
    ('P',  'university_phone',     'University Phone No.',                   20, '82-42-829-7134', '02-460-4227'),
    ('Q',  'entry_date',           'Intended Entry Date (YYYY/MM/DD)',       24, '2026/03/03',     '2026/09/01'),
    ('R',  'korea_address',        'Address in Korea',                       40, '대전서구도안북로88 기숙사', '서울성동구왕십리로222 기숙사'),
    ('S',  'korea_contact',        'Contact No. in Korea',                   22, '+821057171779',  '+82101234567'),
    ('T',  'inviter_name',         'Inviting Organisation Name',             34, 'MOKWON UNIVERSITY', 'HANYANG UNIVERSITY'),
    ('U',  'inviter_reg_no',       'Inviter Business Reg. No.',              22, '305-82-01138',   '206-82-00400'),
    ('V',  'inviter_address',      'Inviter Address',                        40, '대전광역시 서구 도안북로 88', '서울특별시 성동구 왕십리로 222'),
    ('W',  'inviter_phone',        'Inviter Phone No.',                      20, '82-42-829-7134', '02-460-4227'),
    ('X',  'travel_costs',         'Estimated Travel Costs (USD)',           24, '22,000$',        '18,000$'),
    ('Y',  'home_address',         'Home Country Address',                   40, 'DHAKA, BANGLADESH', 'CHITTAGONG, BANGLADESH'),
    ('Z',  'current_address',      'Current Address (if different)',         36, '',               ''),
    ('AA', 'phone',                'Applicant Cell Phone No.',               24, '+8801712345678', '+8801987654321'),
    ('AB', 'email',                        'Applicant Email',                        30, 'shiam@email.com',  'rahman@email.com'),
    ('AC', 'emergency_contact_name',       'Emergency Contact Full Name',            30, 'FATHER NAME',      'MOTHER NAME'),
    ('AD', 'emergency_contact_relationship','Emergency Contact Relationship',        24, 'FATHER',           'MOTHER'),
    ('AE', 'photo_path',                   'Photo File Path (optional)',             40, '',                 ''),
    ('AF', 'status_of_stay',               'Status of Stay (e.g. D-4-1)',           18, 'D-4-1',            'D-4-1'),
    ('AG', 'education',                    'Education (Master/Bachelor/High School/Others)', 28, 'Bachelor', 'Master'),
    ('AH', 'sponsor_name',                 'Sponsor / Payer Name',                  30, 'FATHER NAME',      'AGENCY NAME'),
    ('AI', 'sponsor_relationship',         'Sponsor Relationship to Applicant',     28, 'FATHER',           'AGENCY'),
    ('AJ', 'sponsor_contact',              'Sponsor Contact No.',                   24, '+8801712345678',   '+8801987654321'),
    ('AK', 'school_name',                 'School Name (6.2)',                      30, 'DHAKA COLLEGE',    'RAJSHAHI COLLEGE'),
    ('AL', 'school_location',             'School Location - City/Country (6.3)',   36, 'DHAKA, BANGLADESH','RAJSHAHI, BANGLADESH'),
    # NOC-specific columns
    ('AM', 'noc_father_name',             'NOC - Father Full Name',                 30, 'ANWAR HOSSAIN',    'KARIM UDDIN'),
    ('AN', 'noc_mother_name',             'NOC - Mother Full Name',                 30, 'TANIA AKTER',      'RASHIDA BEGUM'),
]


def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Applicants'

    # ── Header row ─────────────────────────────────────────────────────────
    header_fill  = PatternFill('solid', fgColor='003478')   # Korean flag blue
    accent_fill  = PatternFill('solid', fgColor='C60C30')   # Korean flag red
    header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for i, (col_letter, key, label, width, *_) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = thin_border()

    ws.row_dimensions[1].height = 40

    # ── Sample data rows ───────────────────────────────────────────────────
    data_font  = Font(name='Calibri', size=10)
    data_align = Alignment(horizontal='left', vertical='center')

    for row_idx in range(1, 3):  # 2 sample rows
        r = row_idx + 1
        for c_idx, (col_letter, key, label, width, ex1, ex2) in enumerate(COLUMNS, start=1):
            val = ex1 if row_idx == 1 else ex2
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = thin_border()

    # ── Column widths ──────────────────────────────────────────────────────
    for i, (col_letter, _, _, width, *_) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'

    # ── Education dropdown validation (col AG) ─────────────────────────────
    dv = DataValidation(
        type='list',
        formula1='"Master,Bachelor,High School,Others"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.sqref = 'AG2:AG200'
    ws.add_data_validation(dv)

    # ── Instructions sheet ─────────────────────────────────────────────────
    ins = wb.create_sheet('Instructions')
    ins_rows = [
        ('Korea D-4 Student Visa — Excel Input Guide', True, 13),
        ('', False, 11),
        ('RULES:', True, 11),
        ('1.  Fill one row per applicant starting from row 2.', False, 11),
        ('2.  Do NOT rename or reorder column headers.', False, 11),
        ('3.  Dates: YYYY-MM-DD for issue/expiry, YYYY/MM/DD for entry date.', False, 11),
        ('4.  Birth year/month/day are SEPARATE columns (D, E, F).', False, 11),
        ('5.  Gender: M for Male, F for Female.', False, 11),
        ('6.  Education (col AG): use dropdown — Master / Bachelor / High School / Others', False, 11),
        ('7.  Photo Path (col AE): full absolute path to a JPG/PNG file. Leave blank to skip.', False, 11),
        ('8.  Period of Stay is always 365 DAYS (hardcoded — no column needed).', False, 11),
        ('9.  Save as .xlsx before uploading to the app.', False, 11),
        ('', False, 11),
        ('STATIC FIELDS (always the same — no Excel column needed):', True, 11),
        ('  • Period of Stay → 365 DAYS', False, 11),
        ('  • Nationality / Country of Birth / Country of Passport → BANGLADESHI / BANGLADESH', False, 11),
        ('  • Passport Type → Regular ✓', False, 11),
        ('  • Marital Status → Single ✓', False, 11),
        ('  • Occupation → Student ✓', False, 11),
        ('  • Relationship to Inviter → ADMITTED STUDENT', False, 11),
        ('  • Section 11 Assistance → HANGEUL KOREAN LANGUAGE AND VISA / AGENCY', False, 11),
        ('', False, 11),
        ('Licensed To: Rashid MD Mamun Ur, CEO', True, 11),
    ]
    for row_i, (text, bold, size) in enumerate(ins_rows, start=1):
        cell = ins.cell(row=row_i, column=1, value=text)
        cell.font = Font(bold=bold, size=size)
    ins.column_dimensions['A'].width = 90

    wb.save(OUTPUT_PATH)
    print(f'Excel template saved: {OUTPUT_PATH}')
    print(f'Columns: {len(COLUMNS)}   Sample rows: 2')


if __name__ == '__main__':
    main()
